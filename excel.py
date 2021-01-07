import os
import glob
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import numbers
import win32com.client # win32comをインポートするだけでは上手くいかないので注意！！

def copy_location_number(input_excel_path, input_sheet, col_num, start_row):

    # 正規表現のファイル名から一致するファイルを取得する
    file_list = []
    file_list = glob.glob(os.getcwd() + '/' + input_excel_path)

    workbook = openpyxl.load_workbook(file_list[0], read_only=True)
    sheet = workbook[input_sheet]

    # 最終行を取得
    # openpyxlでは動作が安定しない（指定したシートとは別のシートの最終行をとってきてしまう）ため、win32comで取得）
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(file_list[0], UpdateLinks=0, ReadOnly=True)
    tmp_sheet = workbook.WorkSheets(input_sheet)
    xlUp = -4162
    lastrow = tmp_sheet.Cells(tmp_sheet.Rows.Count, 2).End(xlUp).Row
    workbook.Close()
    workbook = None

    localnum_list = []

    # 地点番号が記載してある最終行までコピーする
    #for i in range(start_row, sheet.max_row + 1):
    for i in range(start_row, lastrow + 1):
        cell_value = sheet.cell(row=i, column=col_num).value

        if cell_value is None:
            continue

        # 数字か判定
        if not cell_value.isdecimal():
            # 数字でない場合
            continue

        # 数字が22桁か判定
        if not len(cell_value) == 22:
            # 22桁でない場合
            continue

        # 重複は追加しない
        #if cell_value not in localnum_list:
        #    localnum_list.append(cell_value)

        localnum_list.append(cell_value)

    return localnum_list

def paste_location_number(localnum_list, input_excel_path, input_sheet, EXCEL_TMP_FILENAME):
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook[input_sheet]

    # 貼り付けする行数とB列の行数を比較する
    blank_line = len(sheet['B']) - len(localnum_list)
    if blank_line > 0:
        # 余分な行を削除
        sheet.delete_rows(idx=len(localnum_list) + 2, amount=blank_line - 1)
    elif blank_line < 0:
        # 行を追加
        sheet.insert_rows(idx=2, amount=-blank_line)
        # 追加した行に元あった行の値と書式を貼り付け
        for i in range(2, -blank_line + 2):
            sheet.cell(row=i, column=1).value = sheet.cell(row=2 -blank_line, column=1).value
            sheet.cell(row=i, column=1)._style = sheet.cell(row=2 -blank_line, column=1)._style
            sheet.cell(row=i, column=2).value = sheet.cell(row=2 -blank_line, column=2).value
            sheet.cell(row=i, column=2)._style = sheet.cell(row=2 -blank_line, column=2)._style
            sheet.cell(row=i, column=3).value = sheet.cell(row=2 -blank_line, column=3).value
            sheet.cell(row=i, column=3)._style = sheet.cell(row=2 -blank_line, column=3)._style
    # 値の貼り付け
    for i in range(2, len(localnum_list) + 2):
        sheet.cell(row=i, column=1).value = localnum_list[i - 2]
        sheet.cell(row=i, column=1).number_format = numbers.FORMAT_NUMBER

    # 罫線を引く
    side = Side(style='thin', color='000000')
    # set border (black thin line)
    border = Border(top=side, bottom=side, left=side, right=side)
    for row in sheet:
        for cell in row:
            if sheet[cell.coordinate].value:
                sheet[cell.coordinate].border = border

    # 印刷範囲を設定する
    area = 'A1:C' + str(len(sheet['B']))
    sheet.print_area = area

    workbook.save(filename = EXCEL_TMP_FILENAME)

def excel_to_pdf(EXCEL_TMP_FILENAME, input_sheet, output_pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    file = excel.Workbooks.Open(os.getcwd() + '/' + EXCEL_TMP_FILENAME, UpdateLinks=0, ReadOnly=True)
    file.WorkSheets[input_sheet].Select()
    outpath = os.getcwd() + '/' + output_pdf_path
    file.ActiveSheet.ExportAsFixedFormat(0, outpath)
    file.Close()
    file = None
